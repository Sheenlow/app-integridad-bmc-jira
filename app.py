import streamlit as st
import pandas as pd
from io import BytesIO


# =============================================================================
# CONSTANTES: Nombres esperados de columnas (referencia por nombre, no por indice)
# Modificar aqui si los nombres de columna cambian en los reportes fuente.
# =============================================================================

# BMC - Work Orders
COL_WO_ID = "WorkOrderID"
COL_WO_SUMMARY = "Summary"
COL_WO_STATUS = "Status"
COL_WO_PRIORITY = "Priority"

# BMC - Problemas (PBI)
COL_PBI_ID = "ProblemID"
COL_PBI_SUMMARY = "Summary"
COL_PBI_STATUS = "Status"

# Jira
COL_JIRA_KEY = "Key"
COL_JIRA_SUMMARY = "Summary"
COL_JIRA_STATUS = "Status"
COL_JIRA_ISSUE_TYPE = "IssueType"
COL_JIRA_ASSIGNEE = "Persona asignada"

# --- Fase 2: Transformacion y cruce ---

# Mapeo de IDs originales al campo unificado BMC_ID
COL_WO_SOURCE_ID = "ID Propuesta"
COL_PBI_SOURCE_ID = "Problema"
COL_BMC_ID = "BMC_ID"

# Columna de estado (presente en los 3 reportes)
COL_ESTADO = "Estado"
COL_ESTADO_BMC = "Estado_BMC"
COL_ESTADO_JIRA = "Estado_JIRA"

# Nombres alternativos que puede tener la columna de estado en los reportes
COL_ESTADO_CANDIDATOS = ["Estado", "Status", "state", "estado"]

# Trazabilidad de origen BMC
COL_ORIGEN = "Origen_BMC"

# --- Fase 3: Reglas de negocio y exportacion ---

COL_ACCION_SUGERIDA = "Acción Sugerida"
COL_MERGE = "_merge"

# Columnas practicas para el preview de Acciones Sugeridas
COLUMNAS_PREVIEW_ACCIONES = [
    COL_BMC_ID,
    COL_JIRA_KEY,
    COL_ESTADO_BMC,
    COL_ESTADO_JIRA,
    COL_JIRA_ASSIGNEE,
    COL_ACCION_SUGERIDA,
]

EQUIVALENCIAS = {
    # PBI
    "Assigned": ["Backlog", "Por hacer", "Tareas por hacer", "En progreso"],
    "Cancelled": ["Cancelado"],
    "Closed": ["Finalizada", "Listo", "Cancelado"],
    "Completed": ["Finalizada", "Listo"],
    "Draft": ["Backlog", "Listo", "Por hacer", "Tareas por hacer"],
    "Pending": ["Stand By"],
    "Under Investigation": ["En progreso", "Dev Doing", "Tareas por hacer"],
    # WO
    "Abandonado": ["Cancelado"],
    "Análisis Técnico": ["Backlog", "Por hacer", "Tareas por hacer", "En progreso", "Stand By"],
    "Cambio Creado": ["En progreso", "Stand By", "Tareas por hacer"],
    "En ejecución": ["En progreso", "Stand By", "Tareas por hacer"],
    "Finalizada": ["Finalizada", "Listo", "Cancelado"],
    "Pendiente Aprobación del negocio": ["Stand By", "Tareas por hacer", "En progreso"],
    "registrado": ["Backlog"],
    "Asignado": ["Backlog"],
}

# --- Validacion Epicas vs Tareas ---

COL_TIPO_INCIDENCIA = "Tipo de Incidencia"
COL_CLAVE_JIRA = "Clave"
COL_PARENT = "parent"
COL_TAREAS_TOTAL = "Tareas Totales"
COL_TAREAS_ABIERTAS = "Tareas Abiertas"
COL_TAREAS_CERRADAS = "Tareas Cerradas"
COL_ESTADOS_TAREAS = "Estados Tareas"
COL_VALIDACION_EPICAS = "Validacion Epica"
COL_CANT_SPRINTS = "Cantidad de Sprints"
COL_CELULA = "Célula"

COL_SPRINT_CANDIDATOS = ["Sprint", "Sprints", "Sprint ID", "Iteracion"]

ESTADOS_FINALES_EPICA = [
    "Done", "Closed", "Finalizada", "Finalizado",
    "Cancelado", "Cancelada", "Cancelled", "Listo",
]


# =============================================================================
# CONFIGURACION DE PAGINA
# =============================================================================

st.set_page_config(
    page_title="BMC ↔ Jira | Conciliacion",
    page_icon="\U0001F310",
    layout="wide",
    initial_sidebar_state="expanded",
)

# --- Estilos CSS minimalista-tech ---
st.markdown(
    """
<style>
    :root {
        --slate-50:  #f8fafc;
        --slate-100: #f1f5f9;
        --slate-200: #e2e8f0;
        --slate-300: #cbd5e1;
        --slate-400: #94a3b8;
        --slate-500: #64748b;
        --slate-600: #475569;
        --slate-700: #334155;
        --slate-800: #1e293b;
        --slate-900: #0f172a;
        --emerald:   #10b981;
        --amber:     #f59e0b;
        --red:       #ef4444;
        --blue:      #3b82f6;
    }

    /* --- Global --- */
    .main .block-container {
        padding-top: 1.5rem;
        padding-bottom: 1rem;
        max-width: 1400px;
    }

    .stApp {
        background: var(--slate-50);
    }

    /* --- Sidebar --- */
    [data-testid="stSidebar"] {
        background: #ffffff;
        border-right: 1px solid var(--slate-200);
    }
    [data-testid="stSidebar"] .block-container {
        padding-top: 1rem;
    }

    /* --- Tabs --- */
    .stTabs [data-baseweb="tab-list"] {
        gap: 0.25rem;
        border-bottom: 1px solid var(--slate-200);
        padding-bottom: 0;
    }
    .stTabs [data-baseweb="tab"] {
        height: 44px;
        font-size: 15px;
        font-weight: 500;
        color: var(--slate-500);
        border-radius: 8px 8px 0 0;
        padding: 0 20px;
    }
    .stTabs [data-baseweb="tab"][aria-selected="true"] {
        color: var(--slate-800);
        background: #ffffff;
        border: 1px solid var(--slate-200);
        border-bottom-color: #ffffff;
    }

    /* --- Metrics --- */
    [data-testid="stMetric"] {
        background: #ffffff;
        border: 1px solid var(--slate-200);
        border-radius: 10px;
        padding: 1rem;
        box-shadow: 0 1px 2px rgba(0,0,0,0.04);
    }
    [data-testid="stMetric"] label {
        font-size: 13px;
        color: var(--slate-500) !important;
    }
    [data-testid="stMetricValue"] {
        font-size: 28px;
        font-weight: 700;
        color: var(--slate-800);
    }

    /* --- Dataframes --- */
    [data-testid="stDataFrame"] {
        border: 1px solid var(--slate-200);
        border-radius: 10px;
        overflow: hidden;
    }

    /* --- Buttons --- */
    .stDownloadButton button {
        background: var(--blue) !important;
        color: #ffffff !important;
        border: none !important;
        border-radius: 8px !important;
        font-weight: 500 !important;
        padding: 0.5rem 1.25rem !important;
        transition: all 0.15s;
    }
    .stDownloadButton button:hover {
        opacity: 0.9;
    }

    /* --- File uploaders --- */
    [data-testid="stFileUploader"] section {
        border-radius: 8px;
        border: 1px dashed var(--slate-300);
    }

    /* --- Containers (border=True cards) --- */
    [data-testid="stNotification"] {
        border-radius: 10px;
    }

    /* --- Hide Streamlit branding --- */
    #MainMenu, footer, header[data-testid="stHeader"] {
        display: none;
    }

    /* --- Info / Warning spacing --- */
    .stAlert {
        border-radius: 8px;
    }
</style>
""",
    unsafe_allow_html=True,
)

# =============================================================================
# INICIALIZACION DE SESSION STATE
# =============================================================================

for key in [
    "df_bmc_wo", "df_bmc_pbi", "df_jira",
    "df_bmc_total", "df_merge", "df_resultado",
    "df_epicas", "df_tareas",
    "df_epicas_filt", "df_tareas_filt", "df_tareas_agg",
    "df_epic_merge", "df_epic_resultado",
]:
    if key not in st.session_state:
        st.session_state[key] = None


# =============================================================================
# FUNCIONES AUXILIARES
# =============================================================================

def detectar_columnas_duplicadas(df: pd.DataFrame, label: str) -> None:
    """Advierte si el DataFrame tiene nombres de columna duplicados."""
    cols = df.columns.tolist()
    duplicados = [c for c in cols if cols.count(c) > 1]
    if duplicados:
        st.warning(
            f"**{label}** - Se detectaron columnas duplicadas: "
            f"{', '.join(sorted(set(duplicados)))}. "
            "Pandas las renombrara con sufijos (.1, .2). "
            "Verifica el archivo fuente."
        )


def listar_hojas_excel(archivo: bytes) -> list[str]:
    """Devuelve los nombres de las hojas de un archivo Excel."""
    from openpyxl import load_workbook
    wb = load_workbook(filename=BytesIO(archivo), read_only=True)
    hojas = wb.sheetnames
    wb.close()
    return hojas


def leer_archivo_robusto(
    uploaded_file, sheet_name=0, permitir_fallback: bool = True
) -> pd.DataFrame:
    """
    Lee un archivo en cascada: Excel (openpyxl) -> CSV (auto-detect separator) -> HTML.
    Para archivos xlsx reales (permitir_fallback=False), si openpyxl falla, intenta
    extraer los datos directamente del ZIP antes de rendirse.
    Reinicia el puntero con seek(0) antes de cada intento.
    Lanza ValueError si todos los formatos fallan.
    """
    uploaded_file.seek(0)
    try:
        return pd.read_excel(uploaded_file, sheet_name=sheet_name, engine="openpyxl")
    except Exception:
        if not permitir_fallback:
            uploaded_file.seek(0)
            try:
                return _extraer_datos_xlsx_desde_zip(uploaded_file, sheet_name)
            except Exception:
                raise

    for enc in ["utf-8", "latin-1", "cp1252"]:
        try:
            uploaded_file.seek(0)
            return pd.read_csv(uploaded_file, sep=None, engine="python", encoding=enc)
        except Exception:
            continue

    try:
        uploaded_file.seek(0)
        return pd.read_html(uploaded_file)[0]
    except Exception:
        pass

    raise ValueError(
        "No se pudo leer el archivo en ningun formato (excel, csv, html)."
    )


def _extraer_datos_xlsx_desde_zip(
    uploaded_file, sheet_name=0
) -> pd.DataFrame:
    """
    Extrae datos de un archivo xlsx corrupto leyendo directamente los XML
    dentro del ZIP, evitando la hoja de estilos que causa el error de openpyxl.
    Soporta sheet_name como indice (int) o nombre (str).
    """
    import zipfile
    from xml.etree import ElementTree as ET

    NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    NS_R = "http://schemas.openxmlformats.org/package/2006/relationships"

    with zipfile.ZipFile(uploaded_file) as zf:

        # --- Shared strings (soporta texto simple y rich text) ---
        strings = []
        if "xl/sharedStrings.xml" in zf.namelist():
            tree = ET.parse(zf.open("xl/sharedStrings.xml"))
            root = tree.getroot()
            for si in root.findall(f"{{{NS}}}si"):
                t = si.find(f"{{{NS}}}t")
                if t is not None and t.text:
                    strings.append(t.text)
                else:
                    # Rich text: concatenar todos los <r><t> dentro de <si>
                    textos = []
                    for r_elem in si.findall(f"{{{NS}}}r"):
                        rt = r_elem.find(f"{{{NS}}}t")
                        if rt is not None and rt.text:
                            textos.append(rt.text)
                    strings.append("".join(textos))

        # --- Resolver hoja por nombre via relaciones del workbook ---
        sheet_file = None

        # Mapa rId -> target desde workbook.xml.rels
        rid_targets = {}
        rels_path = "xl/_rels/workbook.xml.rels"
        if rels_path in zf.namelist():
            rels_tree = ET.parse(zf.open(rels_path))
            for rel in rels_tree.getroot().findall(f"{{{NS_R}}}Relationship"):
                rid = rel.get("Id")
                target = rel.get("Target")
                if rid and target:
                    rid_targets[rid] = target

        # Leer workbook.xml para obtener el target de la hoja deseada
        wb_tree = ET.parse(zf.open("xl/workbook.xml"))
        wb_root = wb_tree.getroot()
        sheets_elem = wb_root.find(f"{{{NS}}}sheets")

        if isinstance(sheet_name, str):
            for s in sheets_elem.findall(f"{{{NS}}}sheet"):
                if s.get("name") == sheet_name:
                    rid = s.get(f"{{{NS_R}}}id") or s.get("r:id")
                    target = rid_targets.get(rid, f"worksheets/sheet{s.get('sheetId', '1')}.xml")
                    sheet_file = f"xl/{target}"
                    break
            if sheet_file is None:
                # Fallback: intentar con el indice 0
                sheet_file = "xl/worksheets/sheet1.xml"
        else:
            idx = int(sheet_name) if sheet_name else 0
            sheet_targets = []
            for s in sheets_elem.findall(f"{{{NS}}}sheet"):
                rid = s.get(f"{{{NS_R}}}id") or s.get("r:id")
                target = rid_targets.get(rid, f"worksheets/sheet{s.get('sheetId', '1')}.xml")
                sheet_targets.append(target)
            if 0 <= idx < len(sheet_targets):
                sheet_file = f"xl/{sheet_targets[idx]}"
            else:
                sheet_file = f"xl/worksheets/sheet{idx + 1}.xml"

        if sheet_file not in zf.namelist():
            disponibles = [n for n in zf.namelist() if "sheet" in n.lower()]
            raise FileNotFoundError(
                f"Hoja '{sheet_file}' no encontrada. "
                f"Hojas disponibles en el ZIP: {disponibles or zf.namelist()[:10]}"
            )

        # --- Parsear la hoja ---
        sheet_tree = ET.parse(zf.open(sheet_file))
        sheet_root = sheet_tree.getroot()
        sheet_data = sheet_root.find(f"{{{NS}}}sheetData")

        if sheet_data is None:
            raise ValueError(
                f"La hoja '{sheet_file}' no contiene <sheetData>."
            )

        rows_data = []
        for row_elem in sheet_data.findall(f"{{{NS}}}row"):
            row = {}
            for cell in row_elem.findall(f"{{{NS}}}c"):
                ref = cell.get("r") or ""
                col_letter = "".join(c for c in ref if c.isalpha())
                value_elem = cell.find(f"{{{NS}}}v")
                formula_elem = cell.find(f"{{{NS}}}f")
                cell_type = cell.get("t", "")

                if value_elem is not None and value_elem.text is not None:
                    if cell_type == "s":
                        idx_str = int(value_elem.text)
                        row[col_letter] = strings[idx_str] if idx_str < len(strings) else ""
                    else:
                        row[col_letter] = value_elem.text
                elif cell_type == "inlineStr":
                    is_elem = cell.find(f"{{{NS}}}is")
                    if is_elem is not None:
                        t_inline = is_elem.find(f"{{{NS}}}t")
                        row[col_letter] = t_inline.text if t_inline is not None and t_inline.text else ""
                    else:
                        row[col_letter] = ""
                elif formula_elem is not None and formula_elem.text:
                    row[col_letter] = formula_elem.text
                else:
                    row[col_letter] = None

            if row:
                rows_data.append(row)

    if not rows_data:
        raise ValueError(
            "No se encontraron datos en el archivo xlsx (extraccion ZIP)."
        )

    df = pd.DataFrame(rows_data)

    if not df.empty:
        primera = df.iloc[0].tolist()
        tiene_headers = any(
            isinstance(v, str) and v.strip() for v in primera if v is not None
        )
        if tiene_headers:
            nuevos = []
            for i, val in enumerate(primera):
                if isinstance(val, str) and val.strip():
                    nuevos.append(val.strip())
                else:
                    nuevos.append(f"Col_{i + 1}")
            df.columns = nuevos
            df = df.iloc[1:].reset_index(drop=True)
        else:
            df.columns = [f"Col_{i + 1}" for i in range(len(df.columns))]

    df = df.fillna("")

    return df


def leer_archivo_subido(archivo, label: str, hoja_preferida: str | None = None) -> pd.DataFrame | None:
    """
    Lee un archivo .csv o .xlsx subido via st.file_uploader.
    - Usa leer_archivo_robusto() con fallback Excel -> CSV -> HTML.
    - Para archivos .xlsx reales, permite elegir hoja si hay mas de una.
    - Si se especifica hoja_preferida y existe, se selecciona automaticamente.
    - Detecta columnas duplicadas y emite advertencia.
    """
    if archivo is None:
        return None

    try:
        nombre = archivo.name.lower()

        if nombre.endswith((".csv", ".xlsx")):
            hoja = 0
            es_fallback = True

            if nombre.endswith(".xlsx"):
                archivo.seek(0)
                magic = archivo.read(4)
                archivo.seek(0)
                es_zip = (magic[:2] == b"PK")

                if es_zip:
                    es_fallback = False
                    try:
                        bytes_archivo = archivo.getvalue()
                        hojas = listar_hojas_excel(bytes_archivo)
                        if hoja_preferida and hoja_preferida in hojas:
                            hoja = hoja_preferida
                        elif len(hojas) == 1:
                            hoja = hojas[0]
                        else:
                            hoja = st.selectbox(
                                f"Hoja a leer para **{label}**",
                                options=hojas,
                                key=f"sheet_{label}",
                            )
                    except Exception:
                        hoja = 0

            with st.spinner(f"Leyendo {label}..."):
                df = leer_archivo_robusto(
                    archivo, sheet_name=hoja, permitir_fallback=es_fallback
                )
        else:
            st.error(
                f"**{label}** - Formato no soportado: '{archivo.name}'. "
                "Usa archivos .csv o .xlsx."
            )
            return None

        if df.empty:
            st.error(f"**{label}** - El archivo esta vacio (0 filas).")
            return None

        detectar_columnas_duplicadas(df, label)

        st.toast(f"✅ {label} cargado — {df.shape[0]:,} filas", icon="✅")
        return df

    except Exception as e:
        st.toast(f"❌ {label} — Error de lectura", icon="❌")
        st.error(f"**{label}** - Error al leer el archivo: {e}")
        return None


def mostrar_resumen(df: pd.DataFrame, label: str) -> None:
    """Muestra shape, columnas y primeras filas de un DataFrame."""
    if df is None or df.empty:
        return

    st.subheader(f"{label}")
    shape_col, cols_col = st.columns([1, 3])
    shape_col.metric("Dimensiones", f"{df.shape[0]:,} filas x {df.shape[1]:,} columnas")
    with cols_col.expander("Columnas detectadas"):
        st.write(df.columns.tolist())

    st.dataframe(df.head(5), use_container_width=True)
    st.markdown("---")


# =============================================================================
# FASE 2: TRANSFORMACION Y CRUCE DE DATOS
# =============================================================================

def _resolver_columnas(df: pd.DataFrame, columnas_deseadas: list[str]) -> list[str]:
    """
    Devuelve las columnas de 'columnas_deseadas' que existen en el DataFrame,
    probando con sufijos _BMC / _JIRA si el nombre exacto no se encuentra.
    """
    disponibles = []
    for col in columnas_deseadas:
        if col in df.columns:
            disponibles.append(col)
        elif f"{col}_BMC" in df.columns:
            disponibles.append(f"{col}_BMC")
        elif f"{col}_JIRA" in df.columns:
            disponibles.append(f"{col}_JIRA")

    faltantes = [c for c in columnas_deseadas if c not in disponibles]
    if faltantes:
        with st.expander("\U0001F50D Columnas requeridas vs disponibles", expanded=True):
            st.markdown(
                f"**No encontradas:** {', '.join(faltantes)}\n\n"
                f"**Columnas disponibles ({len(df.columns)}):** "
                f"{', '.join(sorted(df.columns.tolist()))}"
            )

    return disponibles


def _renombrar_columna_estado(df: pd.DataFrame, nombre_destino: str) -> str | None:
    """
    Busca la columna de estado en 'df' probando varios nombres candidatos
    ('Estado', 'Status', 'state', 'estado') y la renombra a 'nombre_destino'.
    Retorna el nombre original encontrado, o None si no se encontro ninguna.
    """
    for candidato in COL_ESTADO_CANDIDATOS:
        if candidato in df.columns:
            df.rename(columns={candidato: nombre_destino}, inplace=True)
            return candidato
    return None


def normalizar_bmc_wo(df: pd.DataFrame) -> pd.DataFrame | None:
    """
    Renombra 'ID Propuesta' a 'BMC_ID' y la columna de estado a 'Estado_BMC'.
    Agrega columna 'Origen_BMC' = 'WO'.
    Retorna None si la columna fuente no existe.
    """
    if COL_WO_SOURCE_ID not in df.columns:
        st.warning(
            f"**BMC WO** - No se encontro la columna '{COL_WO_SOURCE_ID}'. "
            f"Columnas disponibles: {df.columns.tolist()}"
        )
        return None

    df = df.copy()
    df.rename(columns={COL_WO_SOURCE_ID: COL_BMC_ID}, inplace=True)
    df[COL_ORIGEN] = "WO"

    original = _renombrar_columna_estado(df, COL_ESTADO_BMC)
    if original:
        st.toast(f"✅ WO — '{original}' → Estado_BMC", icon="✅")
    else:
        st.info(f"⚠️ WO — No se detecto columna de estado. Columnas: {df.columns.tolist()}")

    return df


def normalizar_bmc_pbi(df: pd.DataFrame) -> pd.DataFrame | None:
    """
    Renombra 'Problema' a 'BMC_ID' y la columna de estado a 'Estado_BMC'.
    Agrega columna 'Origen_BMC' = 'PBI'.
    Retorna None si la columna fuente no existe.
    """
    if COL_PBI_SOURCE_ID not in df.columns:
        st.warning(
            f"**BMC PBI** - No se encontro la columna '{COL_PBI_SOURCE_ID}'. "
            f"Columnas disponibles: {df.columns.tolist()}"
        )
        return None

    df = df.copy()
    df.rename(columns={COL_PBI_SOURCE_ID: COL_BMC_ID}, inplace=True)
    df[COL_ORIGEN] = "PBI"

    original = _renombrar_columna_estado(df, COL_ESTADO_BMC)
    if original:
        st.toast(f"✅ PBI — '{original}' → Estado_BMC", icon="✅")
    else:
        st.info(f"⚠️ PBI — No se detecto columna de estado. Columnas: {df.columns.tolist()}")

    return df


def unificar_bmc(
    df_wo: pd.DataFrame | None,
    df_pbi: pd.DataFrame | None,
) -> pd.DataFrame | None:
    """
    Concatena los DataFrames normalizados de WO y PBI.
    Soporta que solo uno este disponible.
    """
    partes = []
    if df_wo is not None and not df_wo.empty:
        partes.append(df_wo)
    if df_pbi is not None and not df_pbi.empty:
        partes.append(df_pbi)

    if not partes:
        return None

    df_total = pd.concat(partes, ignore_index=True, sort=False)
    return df_total


def normalizar_jira(df: pd.DataFrame) -> pd.DataFrame | None:
    """
    Renombra la columna de estado a 'Estado_JIRA'.
    Retorna None si el DataFrame es invalido.
    """
    if df is None or df.empty:
        return None

    df = df.copy()

    original = _renombrar_columna_estado(df, COL_ESTADO_JIRA)
    if original:
        st.toast(f"✅ Jira — '{original}' → Estado_JIRA", icon="✅")
    else:
        st.info(f"⚠️ Jira — No se detecto columna de estado. Columnas: {df.columns.tolist()}")

    return df


def cruzar_bmc_jira(
    df_bmc: pd.DataFrame,
    df_jira: pd.DataFrame,
) -> pd.DataFrame | None:
    """
    Outer merge entre BMC y Jira usando 'BMC_ID' como llave.
    indicator=True para trazabilidad de origen de cada registro.
    """
    if df_bmc is None or df_jira is None:
        return None
    if df_bmc.empty or df_jira.empty:
        return None

    try:
        df_merged = pd.merge(
            df_bmc,
            df_jira,
            on=COL_BMC_ID,
            how="outer",
            indicator=True,
            suffixes=("_BMC", "_JIRA"),
        )
        return df_merged

    except Exception as e:
        st.error(f"Error al cruzar BMC con Jira: {e}")
        return None


# =============================================================================
# FASE 3: REGLAS DE NEGOCIO
# =============================================================================

def aplicar_reglas_negocio(df_merge: pd.DataFrame) -> pd.DataFrame | None:
    """
    Agrega la columna 'Accion Sugerida' al DataFrame del merge,
    aplicando reglas de conciliacion fila por fila:
      - left_only  → Falta en Jira - Crear
      - right_only → Sobra en Jira - Revisar/Eliminar
      - both       → comparar Estado_BMC vs Estado_JIRA segun EQUIVALENCIAS
    """
    if df_merge is None or df_merge.empty:
        return None

    try:
        df = df_merge.copy()
        acciones = []

        for _, fila in df.iterrows():
            merge_val = fila.get(COL_MERGE, "")

            if merge_val == "left_only":
                acciones.append("Falta en Jira - Crear")

            elif merge_val == "right_only":
                acciones.append("Sobra en Jira - Revisar/Eliminar")

            elif merge_val == "both":
                estado_bmc = fila.get(COL_ESTADO_BMC)
                estado_jira = fila.get(COL_ESTADO_JIRA)

                if pd.isna(estado_bmc):
                    acciones.append("Actualizar estado: BMC sin estado registrado")
                elif pd.isna(estado_jira):
                    acciones.append(f"Actualizar estado: Jira sin estado, BMC esta en '{estado_bmc}'")
                else:
                    estados_validos = EQUIVALENCIAS.get(estado_bmc)
                    if estados_validos is None:
                        acciones.append(
                            f"Actualizar estado: BMC en '{estado_bmc}', "
                            f"sin equivalencia conocida para '{estado_jira}'"
                        )
                    elif estado_jira in estados_validos:
                        acciones.append("OK - Sincronizado")
                    else:
                        # Buscar si Jira ya avanzo → sugerir avanzar BMC
                        bmc_compatibles = [
                            bmc_s
                            for bmc_s, jira_list in EQUIVALENCIAS.items()
                            if estado_jira in jira_list and bmc_s != estado_bmc
                        ]
                        if bmc_compatibles:
                            sugeridos = ", ".join(bmc_compatibles[:3])
                            acciones.append(
                                f"Revisar: BMC en '{estado_bmc}', "
                                f"Jira ya esta en '{estado_jira}'. "
                                f"Avanzar BMC a: {sugeridos}"
                            )
                        else:
                            lista_str = ", ".join(estados_validos)
                            acciones.append(
                                f"Actualizar estado: BMC en '{estado_bmc}', "
                                f"Jira debe pasar a uno de: [{lista_str}]"
                            )
            else:
                acciones.append("Sin clasificar")

        df[COL_ACCION_SUGERIDA] = acciones
        return df

    except Exception as e:
        st.error(f"Error al aplicar reglas de negocio: {e}")
        return None


# =============================================================================
# VALIDACION EPICAS VS TAREAS
# =============================================================================

def filtrar_epicas(df: pd.DataFrame) -> pd.DataFrame | None:
    """Filtra solo las filas donde Tipo de Incidencia sea Epic."""
    if COL_TIPO_INCIDENCIA not in df.columns:
        st.warning(
            f"No se encontro la columna '{COL_TIPO_INCIDENCIA}'. "
            f"Columnas disponibles: {df.columns.tolist()}"
        )
        return None
    mask = df[COL_TIPO_INCIDENCIA].isin(["Epic", "Épica", "Epica"])
    return df[mask].copy()


def filtrar_tareas(df: pd.DataFrame) -> pd.DataFrame | None:
    """Excluye las epicas, dejando solo tareas e historias."""
    if COL_TIPO_INCIDENCIA not in df.columns:
        st.warning(
            f"No se encontro la columna '{COL_TIPO_INCIDENCIA}'. "
            f"Columnas disponibles: {df.columns.tolist()}"
        )
        return None
    mask = ~df[COL_TIPO_INCIDENCIA].isin(["Epic", "Épica", "Epica"])
    return df[mask].copy()


def agrupar_tareas_por_parent(df_tareas: pd.DataFrame) -> pd.DataFrame | None:
    """
    Agrupa tareas por la columna 'parent' y calcula:
    - total de tareas
    - tareas abiertas (estado no final)
    - tareas cerradas (estado final)
    - lista de estados unicos
    """
    if df_tareas is None or df_tareas.empty:
        return None
    if COL_PARENT not in df_tareas.columns:
        st.warning(f"No se encontro la columna '{COL_PARENT}' en el reporte de Tareas.")
        return None

    # Detectar columna de estado
    if COL_ESTADO not in df_tareas.columns:
        estado_col = None
        for cand in COL_ESTADO_CANDIDATOS:
            if cand in df_tareas.columns:
                estado_col = cand
                break
        if estado_col is None:
            st.warning(
                f"No se encontro columna de estado en Tareas. "
                f"Columnas: {df_tareas.columns.tolist()}"
            )
            return None
    else:
        estado_col = COL_ESTADO

    # Detectar columna de sprint
    sprint_col = None
    for cand in COL_SPRINT_CANDIDATOS:
        if cand in df_tareas.columns:
            sprint_col = cand
            break

    # --- Agregacion de estados ---
    def _agregar_estados(series):
        total = len(series)
        cerradas = series.isin(ESTADOS_FINALES_EPICA).sum()
        abiertas = total - cerradas
        unicos = series.dropna().unique().tolist()
        return pd.Series({
            COL_TAREAS_TOTAL: total,
            COL_TAREAS_ABIERTAS: abiertas,
            COL_TAREAS_CERRADAS: cerradas,
            COL_ESTADOS_TAREAS: ", ".join(sorted(set(str(s) for s in unicos))),
        })

    grouped = df_tareas.groupby(COL_PARENT)[estado_col].apply(_agregar_estados).unstack()
    grouped = grouped.reset_index()
    grouped.rename(columns={COL_PARENT: COL_CLAVE_JIRA}, inplace=True)

    # --- Conteo de sprints unicos ---
    if sprint_col:
        def _contar_sprints(series):
            unicos = series.dropna().unique()
            return len(unicos) if len(unicos) > 0 else 0

        sprints_count = (
            df_tareas.groupby(COL_PARENT)[sprint_col]
            .apply(_contar_sprints)
            .reset_index()
        )
        sprints_count.rename(
            columns={COL_PARENT: COL_CLAVE_JIRA, sprint_col: COL_CANT_SPRINTS},
            inplace=True,
        )
        grouped = pd.merge(grouped, sprints_count, on=COL_CLAVE_JIRA, how="left")
        grouped[COL_CANT_SPRINTS] = grouped[COL_CANT_SPRINTS].fillna(0).astype(int)

    return grouped


def cruzar_epicas_con_tareas(
    df_epicas: pd.DataFrame,
    df_tareas_agg: pd.DataFrame,
) -> pd.DataFrame | None:
    """Outer merge entre epicas y tareas agrupadas, con indicador."""
    if df_epicas is None or df_tareas_agg is None:
        return None
    try:
        merged = pd.merge(
            df_epicas,
            df_tareas_agg,
            on=COL_CLAVE_JIRA,
            how="outer",
            indicator=True,
        )
        return merged
    except Exception as e:
        st.error(f"Error al cruzar Epicas con Tareas: {e}")
        return None


def aplicar_validacion_epicas(df_merged: pd.DataFrame) -> pd.DataFrame | None:
    """
    Agrega columna 'Validacion Epica' al merge Epicas (left) vs Tareas (right):
      - left_only  → Epica sin tareas hijas
      - right_only → Tarea huerfana (sin epica en el reporte)
      - both       → validar consistencia de estados
    """
    if df_merged is None or df_merged.empty:
        return None

    try:
        df = df_merged.copy()
        validaciones = []

        for _, fila in df.iterrows():
            merge_val = fila.get(COL_MERGE, "")

            if merge_val == "left_only":
                validaciones.append("Epica sin tareas hijas")

            elif merge_val == "right_only":
                validaciones.append("Tarea sin epica padre en el reporte")

            elif merge_val == "both":
                estado_epica = fila.get(COL_ESTADO)
                abiertas = fila.get(COL_TAREAS_ABIERTAS, 0)
                cerradas = fila.get(COL_TAREAS_CERRADAS, 0)
                total = fila.get(COL_TAREAS_TOTAL, 0)

                if pd.isna(total) or total == 0:
                    validaciones.append("OK — Sin tareas registradas")
                else:
                    epica_final = (
                        isinstance(estado_epica, str)
                        and estado_epica in ESTADOS_FINALES_EPICA
                    )
                    if epica_final and abiertas > 0:
                        validaciones.append(
                            f"Revisar: Epica cerrada ({estado_epica}) "
                            f"pero tiene {int(abiertas)} tarea(s) abierta(s)"
                        )
                    elif not epica_final and cerradas == total:
                        validaciones.append(
                            f"Todas las tareas cerradas ({int(total)}) "
                            f"— considera cerrar la epica ({estado_epica})"
                        )
                    else:
                        validaciones.append("OK — Consistente")
            else:
                validaciones.append("Sin clasificar")

        df[COL_VALIDACION_EPICAS] = validaciones
        return df

    except Exception as e:
        st.error(f"Error al aplicar validacion de Epicas: {e}")
        return None


# =============================================================================
# FORMATEO DE EXCEL PARA EXPORTACIONES
# =============================================================================

def _formatear_excel(writer, df, sheet_name, columna_color=None):
    """
    Aplica formato al Excel exportado:
    - Auto-ajuste de columnas
    - Header con fondo oscuro y texto blanco
    - Freeze en primera fila
    - Bordes finos
    - Color condicional en la columna especificada
    """
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    ws = writer.sheets[sheet_name]

    # Auto-ajuste de columnas
    for i, col in enumerate(df.columns, 1):
        max_len = max(
            df[col].astype(str).str.len().max(),
            len(str(col)),
        )
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 3, 50)

    # Header
    header_fill = PatternFill("solid", fgColor="1e293b")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    # Freeze
    ws.freeze_panes = "A2"

    # Bordes finos
    thin = Side(style="thin", color="cbd5e1")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            if cell.alignment and not cell.alignment.horizontal:
                cell.alignment = Alignment(vertical="center")

    # Color condicional en columna de estado
    if columna_color and columna_color in df.columns:
        col_idx = list(df.columns).index(columna_color) + 1
        colores = {
            "OK": PatternFill("solid", fgColor="ecfdf5"),
            "Revisar": PatternFill("solid", fgColor="eff6ff"),
            "Falta": PatternFill("solid", fgColor="fef3c7"),
            "Sobra": PatternFill("solid", fgColor="fee2e2"),
            "Actualizar": PatternFill("solid", fgColor="fef2f2"),
            "Epica sin": PatternFill("solid", fgColor="fef3c7"),
            "Tarea sin": PatternFill("solid", fgColor="fef3c7"),
            "Todas": PatternFill("solid", fgColor="eff6ff"),
        }
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            val = str(cell.value or "")
            for prefix, fill in colores.items():
                if val.startswith(prefix):
                    cell.fill = fill
                    break


# =============================================================================
# SIDEBAR: CARGA DE ARCHIVOS
# =============================================================================

# Inicializar variables de uploaders (se asignan dentro del sidebar segun modo)
archivo_bmc_wo = None
archivo_bmc_pbi = None
archivo_jira = None
archivo_epicas = None
archivo_tareas = None

with st.sidebar:
    # Branding
    st.markdown(
        """
        <div style="display:flex;align-items:center;gap:10px;margin-bottom:0.5rem;">
            <div style="background:var(--blue);width:36px;height:36px;border-radius:8px;
                        display:flex;align-items:center;justify-content:center;font-size:18px;">
                \U0001F310
            </div>
            <div>
                <div style="font-weight:700;font-size:15px;color:#0f172a;">BMC ↔ Jira</div>
                <div style="font-size:11px;color:#64748b;">Conciliacion v1.0</div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown("---")

    modo = st.radio(
        "Navegacion",
        ["Conciliacion BMC vs Jira", "Validacion Epicas vs Tareas"],
        key="nav_modo",
        label_visibility="collapsed",
    )

    st.markdown("---")

    # Barras de estado por reporte (compartido entre modos)
    def badge_cargado(label, df_key):
        df = st.session_state.get(df_key)
        if df is not None:
            filas = df.shape[0]
            cols = df.shape[1]
            st.markdown(
                f'<p style="font-size:11px;color:#10b981;margin:0 0 0.25rem 0.5rem;">'
                f'\u2705 {label} — {filas:,} filas, {cols} cols'
                f'</p>',
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                f'<p style="font-size:11px;color:#94a3b8;margin:0 0 0.25rem 0.5rem;">'
                f'\u25CB {label} — Pendiente'
                f'</p>',
                unsafe_allow_html=True,
            )

    if modo == "Conciliacion BMC vs Jira":

        st.markdown(
            '<p style="font-size:12px;font-weight:600;color:#64748b;'
            'text-transform:uppercase;letter-spacing:0.5px;margin-bottom:0.5rem;">'
            '\U0001F517 Descargar reportes</p>',
            unsafe_allow_html=True,
        )

        st.link_button(
            "\U0001F4E4 WO BMC",
            "https://portaltiarcor-or1.onbmc.com/dashboards/d/aecvnwwjfjfggd/wo-panel-general-x-estados?orgId=622379620",
            use_container_width=True,
        )
        st.link_button(
            "\U0001F4E4 PBI BMC",
            "https://portaltiarcor-or1.onbmc.com/dashboards/d/decuhbh2tnri8c/pbi-panel-general-backlog?orgId=622379620",
            use_container_width=True,
        )
        st.link_button(
            "\U0001F4E4 Jira",
            "https://arcor-saic.atlassian.net/issues/?filter=11150",
            use_container_width=True,
        )

        st.markdown("---")

        # --- Reportes ---
        st.markdown(
            '<p style="font-size:12px;font-weight:600;color:#64748b;'
            'text-transform:uppercase;letter-spacing:0.5px;margin-bottom:0.5rem;">'
            '\U0001F4C2 Reportes</p>',
            unsafe_allow_html=True,
        )

        archivo_bmc_wo = st.file_uploader(
            "BMC — Work Orders (WO)",
            type=["csv", "xlsx"],
            key="upload_bmc_wo",
        )

        archivo_bmc_pbi = st.file_uploader(
            "BMC — Problemas (PBI)",
            type=["csv", "xlsx"],
            key="upload_bmc_pbi",
        )

        archivo_jira = st.file_uploader(
            "Jira",
            type=["csv", "xlsx"],
            key="upload_jira",
        )

        badge_cargado("WO", "df_bmc_wo")
        badge_cargado("PBI", "df_bmc_pbi")
        badge_cargado("Jira", "df_jira")

        st.markdown("---")

        cargados = sum(
            1 for key in ["df_bmc_wo", "df_bmc_pbi", "df_jira"]
            if st.session_state.get(key) is not None
        )
        st.progress(cargados / 3, text=f"Progreso: {cargados}/3 reportes")

        # Clean button
        if st.button("\U0001F5D1\uFE0F Limpiar todo", use_container_width=True):
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.rerun()

        st.markdown("---")
        st.caption("Formatos: **.csv** y **.xlsx**  \nColumnas referenciadas por nombre")

    elif modo == "Validacion Epicas vs Tareas":

        st.markdown(
            '<p style="font-size:12px;font-weight:600;color:#64748b;'
            'text-transform:uppercase;letter-spacing:0.5px;margin-bottom:0.5rem;">'
            '\U0001F517 Descargar reportes</p>',
            unsafe_allow_html=True,
        )

        st.link_button(
            "\U0001F4E4 Epicas (Jira)",
            "https://arcor-saic.atlassian.net/issues/?filter=14117",
            use_container_width=True,
        )

        st.link_button(
            "\U0001F4E4 Tareas e Historias (Jira)",
            "https://arcor-saic.atlassian.net/issues/?filter=14116",
            use_container_width=True,
        )

        st.markdown("---")

        st.markdown(
            '<p style="font-size:12px;font-weight:600;color:#64748b;'
            'text-transform:uppercase;letter-spacing:0.5px;margin-bottom:0.5rem;">'
            '\U0001F4C2 Reportes Jira</p>',
            unsafe_allow_html=True,
        )

        archivo_epicas = st.file_uploader(
            "Reporte de Epicas (Jira)",
            type=["csv", "xlsx"],
            key="upload_epicas",
        )

        archivo_tareas = st.file_uploader(
            "Reporte de Tareas e Historias (Jira)",
            type=["csv", "xlsx"],
            key="upload_tareas",
        )

        badge_cargado("Epicas", "df_epicas")
        badge_cargado("Tareas", "df_tareas")

        st.markdown("---")

        cargados_epicas = sum(
            1 for key in ["df_epicas", "df_tareas"]
            if st.session_state.get(key) is not None
        )
        st.progress(cargados_epicas / 2, text=f"Progreso: {cargados_epicas}/2 reportes")

        if st.button("\U0001F5D1\uFE0F Limpiar", use_container_width=True, key="limpiar_epicas"):
            for key in ["df_epicas", "df_tareas", "upload_epicas", "upload_tareas"]:
                if key in st.session_state:
                    del st.session_state[key]
            st.rerun()

        st.markdown("---")
        st.caption("Formatos: **.csv** y **.xlsx**")


# =============================================================================
# AREA PRINCIPAL: PROCESAMIENTO Y VISUALIZACION
# =============================================================================


if modo == "Conciliacion BMC vs Jira":

    # --- Lectura de archivos (se ejecuta siempre) ---

    if archivo_bmc_wo is not None:
        st.session_state.df_bmc_wo = leer_archivo_subido(
            archivo_bmc_wo, "BMC - Work Orders"
        )
    else:
        st.session_state.df_bmc_wo = None

    if archivo_bmc_pbi is not None:
        st.session_state.df_bmc_pbi = leer_archivo_subido(
            archivo_bmc_pbi, "BMC - Problemas"
        )
    else:
        st.session_state.df_bmc_pbi = None

    if archivo_jira is not None:
        st.session_state.df_jira = leer_archivo_subido(archivo_jira, "Jira", hoja_preferida="Your Jira Issues")
    else:
        st.session_state.df_jira = None

    # --- Normalizacion y cruce (se ejecuta siempre) ---

    df_wo_norm = (
        normalizar_bmc_wo(st.session_state.df_bmc_wo)
        if st.session_state.df_bmc_wo is not None
        else None
    )
    df_pbi_norm = (
        normalizar_bmc_pbi(st.session_state.df_bmc_pbi)
        if st.session_state.df_bmc_pbi is not None
        else None
    )
    st.session_state.df_bmc_total = unificar_bmc(df_wo_norm, df_pbi_norm)

    df_jira_norm = normalizar_jira(st.session_state.df_jira)

    if st.session_state.df_bmc_total is not None and df_jira_norm is not None:
        st.session_state.df_merge = cruzar_bmc_jira(
            st.session_state.df_bmc_total, df_jira_norm
        )
    else:
        st.session_state.df_merge = None

    if st.session_state.df_merge is not None:
        with st.spinner("Aplicando reglas de negocio..."):
            st.session_state.df_resultado = aplicar_reglas_negocio(
                st.session_state.df_merge
            )
    else:
        st.session_state.df_resultado = None

    # =============================================================================
    # TABS: UI organizada en 3 pestañas
    # =============================================================================

    tab1, tab2, tab3 = st.tabs([
        "\U0001F4E5 Ingesta de Datos",
        "\U0001F504 Cruce y Reglas de Negocio",
        "\U0001F4CA Resultados y Exportacion",
    ])

    # ── TAB 1: Ingesta ───────────────────────────────────────────────────────

    with tab1:
        col1, col2, col3 = st.columns(3)
        with col1:
            if st.session_state.df_bmc_wo is not None:
                mostrar_resumen(st.session_state.df_bmc_wo, "BMC — Work Orders (WO)")
            else:
                with st.container(border=True):
                    st.markdown(
                        '<p style="color:#94a3b8;text-align:center;padding:2rem 0;">'
                        '\U0001F4E4 Carga el reporte de Work Orders en el sidebar</p>',
                        unsafe_allow_html=True,
                    )

        with col2:
            if st.session_state.df_bmc_pbi is not None:
                mostrar_resumen(st.session_state.df_bmc_pbi, "BMC — Problemas (PBI)")
            else:
                with st.container(border=True):
                    st.markdown(
                        '<p style="color:#94a3b8;text-align:center;padding:2rem 0;">'
                        '\U0001F4E4 Carga el reporte de Problemas en el sidebar</p>',
                        unsafe_allow_html=True,
                    )

        with col3:
            if st.session_state.df_jira is not None:
                mostrar_resumen(st.session_state.df_jira, "Jira")
            else:
                with st.container(border=True):
                    st.markdown(
                        '<p style="color:#94a3b8;text-align:center;padding:2rem 0;">'
                        '\U0001F4E4 Carga el reporte de Jira en el sidebar</p>',
                        unsafe_allow_html=True,
                    )

    # ── TAB 2: Cruce y Reglas ─────────────────────────────────────────────────

    with tab2:
        # BMC Unificado
        if st.session_state.df_bmc_total is not None:
            st.subheader(
                f"\U0001F4E6 BMC Unificado "
                f"({st.session_state.df_bmc_total.shape[0]:,} registros)"
            )
            st.dataframe(
                st.session_state.df_bmc_total.head(10), use_container_width=True
            )
            st.caption(
                "Work Orders (WO) + Problemas (PBI). "
                "Columna `Origen_BMC` indica la fuente."
            )
        else:
            with st.container(border=True):
                st.info(
                    "\U0001F4E4 Carga al menos un reporte BMC (WO o PBI) "
                    "en el sidebar para ver la unificacion."
                )

        # Cruce BMC <-> Jira
        if st.session_state.df_merge is not None:
            st.markdown("---")
            st.subheader("\U0001F91D Cruce BMC ↔ Jira")

            conteo = st.session_state.df_merge[COL_MERGE].value_counts()
            m1, m2, m3 = st.columns(3)
            m1.metric(
                "\U0001F7E2 Sincronizados (both)",
                conteo.get("both", 0),
            )
            m2.metric(
                "\U0001F7E1 Solo BMC (left_only)",
                conteo.get("left_only", 0),
            )
            m3.metric(
                "\U0001F534 Solo Jira (right_only)",
                conteo.get("right_only", 0),
            )

            # Reglas de negocio
            st.markdown("---")
            st.subheader("\U0001F9E0 Acciones Sugeridas")

            def _color_accion(val: str) -> str:
                if val.startswith("OK"):
                    return "background-color: #ecfdf5; color: #065f46;"
                if val.startswith("Revisar"):
                    return "background-color: #eff6ff; color: #1e40af;"
                if val.startswith("Falta"):
                    return "background-color: #fef3c7; color: #92400e;"
                if val.startswith("Sobra"):
                    return "background-color: #fee2e2; color: #991b1b;"
                if val.startswith("Actualizar"):
                    return "background-color: #fef2f2; color: #b91c1c;"
                return ""

            if st.session_state.df_resultado is not None:
                cols_preview = _resolver_columnas(
                    st.session_state.df_resultado, COLUMNAS_PREVIEW_ACCIONES
                )
                df_preview = st.session_state.df_resultado[cols_preview]

                df_no_ok = df_preview[~df_preview[COL_ACCION_SUGERIDA].str.startswith("OK", na=False)]
                if df_no_ok.empty:
                    st.success("Todas las acciones estan OK — sin correcciones requeridas.")
                else:
                    styled = df_no_ok.style.map(
                        _color_accion, subset=[COL_ACCION_SUGERIDA]
                    )
                    st.dataframe(styled, use_container_width=True, height=400)
                    st.caption(
                        f"{df_no_ok.shape[0]:,} registros con acciones requeridas. "
                        "Usa la pestana **Resultados** para el detalle completo."
                    )

                    buf_acciones = BytesIO()
                    with pd.ExcelWriter(buf_acciones, engine="openpyxl") as w:
                        df_no_ok.to_excel(w, index=False, sheet_name="Acciones_Pendientes")
                        _formatear_excel(w, df_no_ok, "Acciones_Pendientes", columna_color=COL_ACCION_SUGERIDA)
                    st.download_button(
                        label="\U0001F4E5 Descargar acciones pendientes (.xlsx)",
                        data=buf_acciones.getvalue(),
                        file_name="acciones_pendientes_conciliacion.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_acciones_tab2",
                    )
        else:
            if st.session_state.df_bmc_total is None:
                pass  # ya se muestra arriba
            elif df_jira_norm is None:
                with st.container(border=True):
                    st.info(
                        "\U0001F4E4 Carga el reporte de Jira en el sidebar "
                        "para realizar el cruce."
                    )

    # ── TAB 3: Resultados y Exportacion ────────────────────────────────────────

    with tab3:
        if st.session_state.df_resultado is not None:
            col_filtro, col_info = st.columns([3, 1])
            with col_filtro:
                acciones_unicas = sorted(
                    st.session_state.df_resultado[COL_ACCION_SUGERIDA].unique()
                )
                seleccion = st.multiselect(
                    "Filtrar por Accion Sugerida",
                    options=acciones_unicas,
                    default=acciones_unicas,
                    key="filtro_accion",
                )

            with col_info:
                with st.popover("\U00002139 Ayuda"):
                    st.markdown(
                        """
                        **Categorias de Accion Sugerida:**
                        - **OK - Sincronizado**: el estado coincide entre sistemas
                        - **Revisar**: BMC esta atrasado — Jira ya avanzo a un estado mas avanzado
                        - **Falta en Jira - Crear**: el registro existe en BMC pero no en Jira
                        - **Sobra en Jira - Revisar/Eliminar**: el registro existe en Jira pero no en BMC
                        - **Actualizar estado**: el estado en Jira no coincide con la equivalencia esperada
                        """
                    )

            if seleccion:
                df_filtrado = st.session_state.df_resultado[
                    st.session_state.df_resultado[COL_ACCION_SUGERIDA].isin(seleccion)
                ]
            else:
                df_filtrado = st.session_state.df_resultado

            total = st.session_state.df_resultado.shape[0]
            mostrados = df_filtrado.shape[0]
            st.caption(f"Mostrando **{mostrados:,}** de **{total:,}** registros")

            # Dataframe estilizado
            def _color_fila_accion(val: str) -> str:
                if val.startswith("OK"):
                    return "background-color: #ecfdf5; color: #065f46;"
                if val.startswith("Revisar"):
                    return "background-color: #eff6ff; color: #1e40af;"
                if val.startswith("Falta"):
                    return "background-color: #fef3c7; color: #92400e;"
                if val.startswith("Sobra"):
                    return "background-color: #fee2e2; color: #991b1b;"
                if val.startswith("Actualizar"):
                    return "background-color: #fef2f2; color: #b91c1c;"
                return ""

            styled_full = df_filtrado.style.map(
                _color_fila_accion, subset=[COL_ACCION_SUGERIDA]
            ).format(precision=0, na_rep="—")

            st.dataframe(
                styled_full,
                use_container_width=True,
                height=520,
                hide_index=True,
            )

            # Descarga
            st.markdown("---")
            buffer = BytesIO()
            with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                df_filtrado.to_excel(writer, index=False, sheet_name="Conciliacion")
                _formatear_excel(writer, df_filtrado, "Conciliacion", columna_color=COL_ACCION_SUGERIDA)

            dl_col, _ = st.columns([1, 3])
            with dl_col:
                st.download_button(
                    label="\U0001F4E5 Descargar resultado_conciliacion.xlsx",
                    data=buffer.getvalue(),
                    file_name="resultado_conciliacion.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
        else:
            with st.container(border=True):
                st.info(
                    "\U0001F4CA Carga los 3 reportes en el sidebar y ve a la pestana "
                    "**Cruce y Reglas** para generar los resultados de conciliacion."
                )


elif modo == "Validacion Epicas vs Tareas":

    st.title("\U0001F4CB Validacion Epicas vs Tareas")
    st.caption("Motor de validacion — Epicas e Historias de Jira")

    # Lectura
    if archivo_epicas is not None:
        st.session_state.df_epicas = leer_archivo_subido(
            archivo_epicas, "Epicas (Jira)"
        )
    else:
        st.session_state.df_epicas = None

    if archivo_tareas is not None:
        st.session_state.df_tareas = leer_archivo_subido(
            archivo_tareas, "Tareas e Historias (Jira)"
        )
    else:
        st.session_state.df_tareas = None

    # Filtrado, agrupacion y cruce
    st.session_state.df_epicas_filt = (
        filtrar_epicas(st.session_state.df_epicas)
        if st.session_state.df_epicas is not None
        else None
    )
    st.session_state.df_tareas_filt = (
        filtrar_tareas(st.session_state.df_tareas)
        if st.session_state.df_tareas is not None
        else None
    )
    st.session_state.df_tareas_agg = agrupar_tareas_por_parent(
        st.session_state.df_tareas_filt
    )

    if (
        st.session_state.df_epicas_filt is not None
        and st.session_state.df_tareas_agg is not None
    ):
        st.session_state.df_epic_merge = cruzar_epicas_con_tareas(
            st.session_state.df_epicas_filt, st.session_state.df_tareas_agg
        )
    else:
        st.session_state.df_epic_merge = None

    if st.session_state.df_epic_merge is not None:
        with st.spinner("Aplicando validacion de Epicas..."):
            st.session_state.df_epic_resultado = aplicar_validacion_epicas(
                st.session_state.df_epic_merge
            )
    else:
        st.session_state.df_epic_resultado = None

    # Tabs
    tab1, tab2, tab3 = st.tabs([
        "\U0001F4E5 Ingesta",
        "\U0001F504 Cruce y Validacion",
        "\U0001F4CA Resultados",
    ])

    with tab1:
        col1, col2 = st.columns(2)
        with col1:
            if st.session_state.df_epicas is not None:
                mostrar_resumen(st.session_state.df_epicas, "Epicas (Jira)")
            else:
                with st.container(border=True):
                    st.markdown(
                        '<p style="color:#94a3b8;text-align:center;padding:2rem 0;">'
                        '\U0001F4E4 Carga el reporte de Epicas en el sidebar</p>',
                        unsafe_allow_html=True,
                    )
            if st.session_state.df_epicas_filt is not None:
                st.caption(
                    f"Epicas filtradas: {st.session_state.df_epicas_filt.shape[0]:,} "
                    f"(original: {st.session_state.df_epicas.shape[0]:,})"
                )

        with col2:
            if st.session_state.df_tareas is not None:
                mostrar_resumen(st.session_state.df_tareas, "Tareas e Historias (Jira)")
            else:
                with st.container(border=True):
                    st.markdown(
                        '<p style="color:#94a3b8;text-align:center;padding:2rem 0;">'
                        '\U0001F4E4 Carga el reporte de Tareas en el sidebar</p>',
                        unsafe_allow_html=True,
                    )
            if st.session_state.df_tareas_filt is not None:
                st.caption(
                    f"Tareas filtradas: {st.session_state.df_tareas_filt.shape[0]:,} "
                    f"(original: {st.session_state.df_tareas.shape[0]:,})"
                )

    with tab2:
        if st.session_state.df_epic_merge is not None:
            conteo = st.session_state.df_epic_merge[COL_MERGE].value_counts()
            m1, m2, m3 = st.columns(3)
            m1.metric(
                "\U0001F7E2 Vinculadas (both)",
                conteo.get("both", 0),
            )
            m2.metric(
                "\U0001F7E1 Solo Tareas (left_only)",
                conteo.get("left_only", 0),
            )
            m3.metric(
                "\U0001F534 Solo Epicas (right_only)",
                conteo.get("right_only", 0),
            )

            st.markdown("---")
            st.subheader("\U0001F9E0 Validacion de Epicas")

            if st.session_state.df_epic_resultado is not None:
                cols_view = [
                    COL_CLAVE_JIRA,
                    COL_CELULA,
                    COL_ESTADO,
                    COL_TAREAS_TOTAL,
                    COL_TAREAS_ABIERTAS,
                    COL_TAREAS_CERRADAS,
                    COL_CANT_SPRINTS,
                    COL_ESTADOS_TAREAS,
                    COL_VALIDACION_EPICAS,
                ]
                cols_ok = [c for c in cols_view if c in st.session_state.df_epic_resultado.columns]
                df_view = st.session_state.df_epic_resultado[cols_ok]

                df_no_ok = df_view[~df_view[COL_VALIDACION_EPICAS].str.startswith("OK", na=False)]
                if df_no_ok.empty:
                    st.success("Todas las epicas estan OK — sin acciones requeridas.")
                else:
                    def _color_validacion(v):
                        if v.startswith("OK"):
                            return "background-color: #ecfdf5; color: #065f46;"
                        if v.startswith("Revisar") or v.startswith("Todas"):
                            return "background-color: #eff6ff; color: #1e40af;"
                        if v.startswith("Epica sin") or v.startswith("Tarea sin"):
                            return "background-color: #fef3c7; color: #92400e;"
                        return ""

                    styled = df_no_ok.style.map(
                        _color_validacion, subset=[COL_VALIDACION_EPICAS]
                    ).format(precision=0, na_rep="—")
                    st.dataframe(styled, use_container_width=True, height=400)
                    st.caption(
                        f"{df_no_ok.shape[0]:,} registros con acciones requeridas. "
                        "Usa la pestana **Resultados** para el detalle completo."
                    )

                    buf_tab2 = BytesIO()
                    with pd.ExcelWriter(buf_tab2, engine="openpyxl") as w:
                        df_no_ok.to_excel(w, index=False, sheet_name="Validacion_Epicas")
                        _formatear_excel(w, df_no_ok, "Validacion_Epicas", columna_color=COL_VALIDACION_EPICAS)
                    st.download_button(
                        label="\U0001F4E5 Descargar solo pendientes (.xlsx)",
                        data=buf_tab2.getvalue(),
                        file_name="validacion_epicas_pendientes.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_epic_tab2",
                    )
        else:
            with st.container(border=True):
                st.info(
                    "\U0001F4E4 Carga ambos reportes (Epicas y Tareas) en el sidebar "
                    "para ver la validacion."
                )

    with tab3:
        if st.session_state.df_epic_resultado is not None:
            vals_unicas = sorted(
                st.session_state.df_epic_resultado[COL_VALIDACION_EPICAS].unique()
            )
            sel = st.multiselect(
                "Filtrar por Validacion de Epica",
                options=vals_unicas,
                default=vals_unicas,
                key="filtro_epic",
            )
            df_filt = (
                st.session_state.df_epic_resultado
                if not sel
                else st.session_state.df_epic_resultado[
                    st.session_state.df_epic_resultado[COL_VALIDACION_EPICAS].isin(sel)
                ]
            )
            st.caption(
                f"Mostrando **{df_filt.shape[0]:,}** de "
                f"**{st.session_state.df_epic_resultado.shape[0]:,}** registros"
            )

            def _color_fila_validacion(v):
                if v.startswith("OK"):
                    return "background-color: #ecfdf5; color: #065f46;"
                if v.startswith("Revisar") or v.startswith("Todas"):
                    return "background-color: #eff6ff; color: #1e40af;"
                if v.startswith("Epica sin") or v.startswith("Tarea sin"):
                    return "background-color: #fef3c7; color: #92400e;"
                return ""

            styled_full = df_filt.style.map(
                _color_fila_validacion, subset=[COL_VALIDACION_EPICAS]
            ).format(precision=0, na_rep="—")

            st.dataframe(
                styled_full, use_container_width=True, height=520, hide_index=True
            )

            st.markdown("---")
            buffer = BytesIO()
            with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                df_filt.to_excel(writer, index=False, sheet_name="Validacion_Epicas")
                _formatear_excel(writer, df_filt, "Validacion_Epicas", columna_color=COL_VALIDACION_EPICAS)

            dl_col, _ = st.columns([1, 3])
            with dl_col:
                st.download_button(
                    label="\U0001F4E5 Descargar validacion_epicas.xlsx",
                    data=buffer.getvalue(),
                    file_name="validacion_epicas.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
        else:
            with st.container(border=True):
                st.info(
                    "\U0001F4CA Carga ambos reportes en el sidebar y ve a la pestana "
                    "**Cruce y Validacion** para generar los resultados."
                )
